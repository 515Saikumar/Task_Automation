# ============================================================
# app.py
# ============================================================

# -----------------------------
# Python standard library
# -----------------------------
import json
from datetime import datetime

# -----------------------------
# FastAPI imports
# -----------------------------
from fastapi import FastAPI, UploadFile, File, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# -----------------------------
# LangChain imports
# -----------------------------
from langchain_groq import ChatGroq
import os
from dotenv import load_dotenv

load_dotenv()

try:
    from langchain.agents import AgentExecutor, create_tool_calling_agent
except ImportError:
    try:
        from langchain.agents.agent import AgentExecutor
        from langchain.agents.tool_calling_agent.base import create_tool_calling_agent
    except ImportError:
        try:
            from langchain_classic.agents import AgentExecutor, create_tool_calling_agent
        except ImportError as e:
            raise ImportError(
                "Could not import AgentExecutor. In newer LangChain versions, "
                "AgentExecutor is moved. Please run `pip install langchain-classic` "
                "or install an older version of LangChain."
            ) from e
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.tools import tool

# -----------------------------
# Existing project imports
# -----------------------------
from api.upload_api import router as upload_router
from api.auth_api import router as auth_router
from api.employee_api import router as employee_router
from api.workflow_api import router as workflow_router
from database.mongodb import workprogress_collection, tasks_collection


# ============================================================
# CREATE FASTAPI APPLICATION
# ============================================================

app = FastAPI(
    title="Task Automation API",
    description="Task Automation and Employee Work Progress API",
    version="1.0.0"
)

# ============================================================
# CORS MIDDLEWARE (FIX FOR BLOCKED REQUESTS / 405 ERRORS)
# ============================================================

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  
    allow_credentials=True,
    allow_methods=["*"],  
    allow_headers=["*"],  
)


# ============================================================
# REGISTER ROUTERS
# ============================================================

app.include_router(upload_router)
app.include_router(auth_router)
app.include_router(employee_router)
app.include_router(workflow_router)


# ============================================================
# GET RECENT EXCEL FILES (FIX FOR /api/tasks 404)
# ============================================================

@app.get("/api/tasks")
def get_recent_task_files():
    """
    Fetches the list of recently uploaded Excel files for the Admin Dashboard.
    Excludes the heavy binary 'file_data' to keep the response fast.
    """
    try:
        files = list(tasks_collection.find({}, {"file_data": 0}))
        
        result = []
        for f in files:
            result.append({
                "_id": str(f["_id"]),
                "filename": f.get("filename", "Unknown File"),
                "status": f.get("status", "Uploaded"),
                "uploaded_at": f.get("uploaded_at", datetime.now().isoformat())
            })
            
        return result
    except Exception as e:
        print(f"Error fetching tasks: {e}")
        return []

# ============================================================
# EXPORT TASKS TO CSV (NEW FEATURE)
# ============================================================
import pandas as pd
from fastapi.responses import StreamingResponse
import io

@app.get("/api/export-tasks")
def export_tasks_excel(time_filter: str = None):
    try:
        from datetime import datetime, timedelta, timezone
        import json
        
        query = {}
        now = datetime.now(timezone.utc)
        
        if time_filter == "today":
            start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
            query = {"$or": [{"updatedAt": {"$gte": start_of_day}}, {"createdAt": {"$gte": start_of_day}}]}
        elif time_filter == "this_week":
            start_of_week = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            query = {"$or": [{"updatedAt": {"$gte": start_of_week}}, {"createdAt": {"$gte": start_of_week}}]}
        elif time_filter == "last_week":
            start_of_this_week = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            start_of_last_week = start_of_this_week - timedelta(days=7)
            query = {"$or": [
                {"updatedAt": {"$gte": start_of_last_week, "$lt": start_of_this_week}},
                {"createdAt": {"$gte": start_of_last_week, "$lt": start_of_this_week}}
            ]}

        results = list(workprogress_collection.find(query, {"_id": 0}))
        
        if not results:
            df = pd.DataFrame([{"Message": "No tasks found for this period."}])
        else:
            # Prepare clean data
            clean_list = []
            for r in results:
                contribs = r.get('contributors', [])
                if isinstance(contribs, list) and contribs:
                    contrib_str = ', '.join([f"{c.get('name')} ({c.get('empid')})" for c in contribs])
                else:
                    contrib_str = ""
                
                clean_list.append({
                    "Employee Name": r.get('empname', ''),
                    "Employee ID": r.get('empid', ''),
                    "Task": r.get('task', ''),
                    "Status": r.get('status', ''),
                    "Priority": r.get('priority', ''),
                    "Category": r.get('category', ''),
                    "Due Date": r.get('duedate', ''),
                    "Remarks": r.get('remarks', ''),
                    "Previous Contributors": contrib_str
                })
            
            # --- LLM RESTRUCTURING ---
            # Ask the LLM to professionally rewrite the 'Task' and 'Remarks' fields.
            llm_prompt = f"""
            You are a professional Executive Assistant. I will provide a JSON array of employee tasks.
            Your job is to rewrite the "Task" and "Remarks" fields to sound highly professional, clear, and well-structured.
            Fix any spelling or grammar mistakes (e.g. "crate a dockar file" -> "Create a Docker file").
            Return ONLY the updated JSON array of objects. Do NOT include markdown code blocks, just the raw JSON array.
            
            Here is the data:
            {json.dumps(clean_list, default=str)}
            """
            
            try:
                # We reuse the llm defined below in app.py
                llm_response = llm.invoke(llm_prompt)
                raw_json = llm_response.content.strip()
                # Clean up if the LLM adds markdown by mistake
                if raw_json.startswith("```json"):
                    raw_json = raw_json[7:-3]
                elif raw_json.startswith("```"):
                    raw_json = raw_json[3:-3]
                
                professional_data = json.loads(raw_json)
                df = pd.DataFrame(professional_data)
            except Exception as e:
                print(f"LLM rewriting failed, falling back to original data: {e}")
                df = pd.DataFrame(clean_list)
            
            # Format dates beautifully if they exist
            if 'Due Date' in df.columns:
                df['Due Date'] = pd.to_datetime(df['Due Date'], errors='ignore').dt.strftime('%Y-%m-%d')
            
        output = io.BytesIO()
        
        # Write to actual Excel using openpyxl for formatting
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Tasks')
            worksheet = writer.sheets['Tasks']
            
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Define styles
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            wrap_alignment = Alignment(wrap_text=True, vertical='top')
            
            # Apply styles to headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
            # Set column widths and apply wrap text to Task and Remarks
            col_widths = {
                "A": 20, # Employee Name
                "B": 15, # Employee ID
                "C": 45, # Task
                "D": 15, # Status
                "E": 12, # Priority
                "F": 15, # Category
                "G": 12, # Due Date
                "H": 45, # Remarks
                "I": 25  # Previous Contributors
            }
            
            for col_letter, width in col_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Additional styles for conditional formatting
            overdue_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") # Light Red
            completed_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid") # Light Green
            
            today_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            
            # Apply wrap text to all data rows and conditional formatting
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = wrap_alignment
                    
                # Get Status and Due Date values
                status_cell = row[3] # Status is 4th column (Index 3)
                due_date_cell = row[6] # Due Date is 7th column (Index 6)
                
                status_val = str(status_cell.value).strip().lower() if status_cell.value else ""
                due_date_val = due_date_cell.value
                
                # Check for Completion (Green)
                if status_val in ["approved", "done", "completed"]:
                    due_date_cell.fill = completed_fill
                    status_cell.fill = completed_fill
                else:
                    # Check for Overdue (Red)
                    if due_date_val:
                        try:
                            # Due Date is currently a string in YYYY-MM-DD format based on our earlier formatting
                            due_date_obj = datetime.strptime(str(due_date_val), "%Y-%m-%d").replace(tzinfo=timezone.utc)
                            if due_date_obj < today_date:
                                due_date_cell.fill = overdue_fill
                                status_cell.fill = overdue_fill
                        except Exception as e:
                            pass # Silently pass if date parsing fails

        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="tasks_professional_{time_filter or "all"}.xlsx"'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return {"error": str(e)}


# ============================================================
# UPLOAD EXCEL FILE (FIX FOR /api/upload-excel 404)
# ============================================================

@app.post("/api/upload-excel")
async def upload_excel_file(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    """
    Receives the Excel file from the React Admin Dashboard, saves it to MongoDB,
    and triggers the AI processing pipeline in the background.
    """
    try:
        # Read the file data as bytes
        file_data = await file.read()
        
        # Create the document for MongoDB
        doc = {
            "filename": file.filename,
            "file_data": file_data,
            "status": "Uploaded",
            "uploaded_at": datetime.now().isoformat()
        }
        
        # Insert into the database
        result = tasks_collection.insert_one(doc)
        file_id = str(result.inserted_id)
        
        # Import your processing function from main.py
        try:
            from main import process_excel
            # Run the AI allocation in the background so the UI doesn't hang
            background_tasks.add_task(process_excel, file_id)
        except ImportError:
            print("⚠️ Could not import process_excel from main.py. Is main.py in the same directory?")
        
        return {"success": True, "file_id": file_id}
        
    except Exception as e:
        print(f"Upload error: {str(e)}")
        return {"success": False, "detail": str(e)}


# ============================================================
# WORK PROGRESS DATABASE TOOL
# ============================================================

@tool
def query_workprogress(search_term: str) -> str:
    """
    Queries employee work progress from MongoDB.

    Input should be a search term such as:
    - employee ID
    - task name
    - status
    """
    try:
        query = {
            "$or": [
                {
                    "empid": {
                        "$regex": search_term,
                        "$options": "i"
                    }
                },
                {
                    "empname": {
                        "$regex": search_term,
                        "$options": "i"
                    }
                },
                {
                    "contributors.empid": {
                        "$regex": search_term,
                        "$options": "i"
                    }
                },
                {
                    "contributors.name": {
                        "$regex": search_term,
                        "$options": "i"
                    }
                },
                {
                    "task": {
                        "$regex": search_term,
                        "$options": "i"
                    }
                },
                {
                    "category": {
                        "$regex": search_term,
                        "$options": "i"
                    }
                },
                {
                    "status": {
                        "$regex": search_term,
                        "$options": "i"
                    }
                }
            ]
        }

        # Query MongoDB
        results = list(
            workprogress_collection.find(
                query,
                {
                    "_id": 0
                }
            )
        )

        # No results
        if not results:
            return (
                f"No work progress found for "
                f"'{search_term}'."
            )

        # Convert datetime values to JSON-compatible strings
        for doc in results:
            for key, value in doc.items():
                if isinstance(value, datetime):
                    doc[key] = value.isoformat()

        # Return results as JSON string
        return json.dumps(
            results,
            default=str
        )

    except Exception as e:
        return (
            f"Error querying database: {str(e)}"
        )


@tool
def get_all_workprogress(time_filter: str = None) -> str:
    """
    Returns employee work progress records. Use this tool when the user asks for:
    - A specific team, domain, or role (e.g. "aiml team", "frontend")
    - Tasks completed "today", "this week", "last week", etc.
    - All completed/pending tasks in general
    - A summary of all tasks
    
    If the user asks for updates within a specific timeframe, pass 'time_filter' as one of: 'today', 'this_week', 'last_week'. Leave blank for all records.
    """
    try:
        from datetime import datetime, timedelta, timezone
        
        query = {}
        now = datetime.now(timezone.utc)
        
        if time_filter == "today":
            start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
            query = {"$or": [{"updatedAt": {"$gte": start_of_day}}, {"createdAt": {"$gte": start_of_day}}]}
        elif time_filter == "this_week":
            # Monday of current week
            start_of_week = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            query = {"$or": [{"updatedAt": {"$gte": start_of_week}}, {"createdAt": {"$gte": start_of_week}}]}
        elif time_filter == "last_week":
            # Monday of last week to Sunday of last week
            start_of_this_week = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0)
            start_of_last_week = start_of_this_week - timedelta(days=7)
            query = {"$or": [
                {"updatedAt": {"$gte": start_of_last_week, "$lt": start_of_this_week}},
                {"createdAt": {"$gte": start_of_last_week, "$lt": start_of_this_week}}
            ]}

        results = list(workprogress_collection.find(query, {"_id": 0}))
        
        # Enrich with employee details so the LLM can easily group by team
        for doc in results:
            empid = doc.get("empid")
            if empid:
                from database.mongodb import emp_collection
                emp = emp_collection.find_one({"employee_id": empid})
                if emp:
                    doc["employee_team"] = emp.get("primary_category", "")
                    doc["employee_department"] = emp.get("department", "")
            
            for key, value in doc.items():
                if isinstance(value, datetime):
                    doc[key] = value.isoformat()
                    
        return json.dumps(results, default=str)
    except Exception as e:
        return f"Error querying database: {str(e)}"

@tool
def allocate_task(empid: str, task: str, priority: str = "Normal", category: str = "General", due_date: str = "") -> str:
    """
    Manually allocates a new task to a specific employee without using the Excel file.
    Use this tool when the user asks to "assign", "allocate", or "give" a new task to an employee.
    Ensure you have their employee ID (empid) before calling this.
    """
    try:
        from database.mongodb import emp_collection, workprogress_collection
        from datetime import datetime, timezone
        
        emp = emp_collection.find_one({"employee_id": empid})
        if not emp:
            # Fallback to searching by name if they passed a name instead of ID
            emp = emp_collection.find_one({"name": {"$regex": empid, "$options": "i"}})
            if not emp:
                return f"Error: Could not find any employee with ID or name matching '{empid}'."
                
        try:
            due_date_val = datetime.strptime(due_date, "%Y-%m-%d").strftime("%Y-%m-%d") if due_date else ""
        except ValueError:
            due_date_val = due_date
            
        workprogress_doc = {
            "empid": emp["employee_id"],
            "empname": emp["name"],
            "task": task,
            "description": "",
            "priority": priority,
            "required_skills": [],
            "category": category,
            "dependencies": [],
            "duedate": due_date_val,
            "taskupdate": "",
            "status": "In Progress",
            "remarks": "",
            "contributors": [],
            "createdAt": datetime.now(timezone.utc),
            "updatedAt": datetime.now(timezone.utc)
        }
        
        workprogress_collection.insert_one(workprogress_doc)
        
        new_active = emp.get("active_tasks", 0) + 1
        max_tasks = emp.get("max_tasks", 4)
        new_availability = new_active < max_tasks
        
        emp_collection.update_one(
            {"employee_id": emp["employee_id"]},
            {"$set": {
                "active_tasks": new_active,
                "availability": new_availability
            }}
        )
        
        return f"Successfully allocated task '{task}' to {emp['name']} ({emp['employee_id']})."
    except Exception as e:
        return f"Failed to allocate task: {str(e)}"

# ============================================================
# INITIALIZE OPENAI LLM
# ============================================================

llm = ChatGroq(
    model_name="openai/gpt-oss-20b",
    temperature=0,
    groq_api_key=os.getenv("GROQ_API_KEY")
)

# ============================================================
# LANGCHAIN TOOLS
# ============================================================

tools = [
    query_workprogress,
    get_all_workprogress,
    allocate_task
]


# ============================================================
# LANGCHAIN PROMPT
# ============================================================

prompt = ChatPromptTemplate.from_messages(
    [
        (
            "system",
            """
You are a helpful HR and Project Management assistant.

The current date and time is: {current_datetime}

You can access employee work progress information
from the MongoDB database using the provided tools.

When the user asks about specific names or IDs:
use the `query_workprogress` tool. If the user makes a typo in the name (e.g. "priys" instead of "Priya"), try to correct it before searching.
If `query_workprogress` returns no results, DO NOT immediately give up. Instead, use the `get_all_workprogress` tool to retrieve all records and manually look for similar employee names to handle typos gracefully.

When the user asks to assign a new task to someone:
use the `allocate_task` tool! You can pass the employee's ID (or name), the task description, and optional due date/priority.

When the user asks about:
- A specific team, domain, or role (e.g. "aiml team", "frontend")
- Tasks completed "today", "this week", "last week", or on a specific date
- All completed/pending tasks in general
- A summary of all tasks
use the `get_all_workprogress` tool! Pull all records and filter the JSON list yourself to find EVERY single task (both completed and ongoing) related to that team/domain or time period. Do not use `query_workprogress` for team/time queries, as it might miss tasks.

Note: Tasks are usually considered completed if their status is 'Done' or 'Approved'.

Answer the user clearly and concisely.

CRITICAL FORMATTING INSTRUCTIONS:
You MUST format your response strictly in HTML.
DO NOT use Markdown (no asterisks, no hash tags, no markdown tables). Do NOT wrap the output in ```html codeblocks. Return pure HTML.

EXPORTING TO EXCEL/CSV:
If the user explicitly asks to "download", "export", or wants the data "into excel" or "CSV", DO NOT print out the raw CSV text. Instead, provide a beautiful HTML download link that points to the `/api/export-tasks` endpoint. 
Example: `<a href="http://127.0.0.1:10000/api/export-tasks?time_filter=this_week" download="Tasks.xlsx" style="display: inline-block; padding: 10px 20px; background-color: #007bff; color: white; text-decoration: none; border-radius: 5px; font-weight: bold;">📥 Download Professional Excel File</a>`
(Make sure to use the correct `time_filter`: 'today', 'this_week', 'last_week', or omit it for all tasks).

When the user asks about a team, domain, role, or asks for "overall updates" (e.g. "aiml team", "frontend tasks", "overall updates"), output a clean HTML <table> with columns: Employee Name, Employee ID, Task, Status, Due Date, and Previous Contributors.
When the user asks about a specific person or their personal work progress (e.g. "what about emp001", "Aarav's tasks"), provide a nicely formatted HTML list (<ul> <li>) and use <strong> for emphasis. 
CRITICAL: When listing tasks for a specific person, you MUST include ALL tasks where they are the current owner (`empid`) AND any tasks where they are listed in the `contributors` array! Do not skip tasks just because they were reassigned to someone else.
IMPORTANT: Tasks might have a 'contributors' array if they were reassigned. If so, you MUST explicitly list the previous contributors (e.g. EMP001) in both the summary tables and personal lists so they get credit for the work they did before reassignment!
Format beautifully and be concise.

Do not invent employee information.

If the database does not contain the requested information,
clearly tell the user that no matching records were found.
"""
        ),
        (
            "placeholder",
            "{chat_history}"
        ),
        (
            "human",
            "{input}"
        ),
        (
            "placeholder",
            "{agent_scratchpad}"
        )
    ]
)


# ============================================================
# CREATE TOOL-CALLING AGENT
# ============================================================

agent = create_tool_calling_agent(
    llm,
    tools,
    prompt
)


# ============================================================
# CREATE AGENT EXECUTOR
# ============================================================

agent_executor = AgentExecutor(
    agent=agent,
    tools=tools,
    verbose=True
)


# ============================================================
# CHAT REQUEST MODEL
# ============================================================

from typing import List

class ChatMessage(BaseModel):
    role: str
    text: str

class ChatRequest(BaseModel):
    message: str
    history: List[ChatMessage] = []


# ============================================================
# CHAT API
# ============================================================

@app.post("/api/chat")
def chat_endpoint(request: ChatRequest):
    try:
        from langchain_core.messages import HumanMessage, AIMessage
        
        # Build memory from frontend history
        chat_history = []
        for msg in request.history:
            if msg.role == 'user':
                chat_history.append(HumanMessage(content=msg.text))
            elif msg.role == 'ai':
                chat_history.append(AIMessage(content=msg.text))

        # Run LangChain agent
        response = agent_executor.invoke(
            {
                "input": request.message,
                "chat_history": chat_history,
                "current_datetime": datetime.now().isoformat()
            }
        )

        return {
            "response": response["output"]
        }

    except Exception as e:
        return {
            "response": (
                f"Sorry, I encountered an error: "
                f"{str(e)}"
            )
        }


# ============================================================
# HEALTH CHECK
# ============================================================

@app.get("/")
def root():
    return {
        "message": "Task Automation API is running",
        "status": "success"
    }


# ============================================================
# API HEALTH CHECK
# ============================================================

@app.get("/health")
def health_check():
    return {
        "status": "healthy"
    }


# ============================================================
# RUN APPLICATION
# ============================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "app:app",
        host="127.0.0.1",
        port=10000, 
        reload=True
    )